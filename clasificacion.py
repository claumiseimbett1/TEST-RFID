#!/usr/bin/env python3
"""
Genera clasificación por tiempo a partir de resultados_con_nadadores.csv:
- CSV y Excel con posición general y por categoría+sexo (simultáneo).
- PDF con tablas: clasificación general y por categoría y sexo.
"""
import csv
import io
import sys

CSV_ENTRADA = "resultados_con_nadadores.csv"
CSV_SALIDA = "clasificacion.csv"
XLSX_SALIDA = "clasificacion.xlsx"
PDF_SALIDA = "clasificacion.pdf"


def _leer_archivo_texto(archivo: str) -> str:
    """Lee el archivo probando utf-8, cp1252 y latin-1."""
    with open(archivo, "rb") as f:
        raw = f.read()
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _pdf_safe(s: str, max_len: int = 50) -> str:
    """Texto seguro para PDF con Helvetica (sin Unicode especial: —, acentos, etc.)."""
    if not s:
        return ""
    t = str(s)[:max_len]
    reemplazos = [
        ("—", "-"), ("–", "-"), ("´", "'"), ("'", "'"), ("'", "'"),
        ("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"),
        ("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U"),
        ("ñ", "n"), ("Ñ", "N"), ("ü", "u"), ("Ü", "U"),
    ]
    for a, b in reemplazos:
        t = t.replace(a, b)
    return "".join(c if ord(c) < 128 else "?" for c in t)


def _parse_tiempo(s: str):
    """Convierte tiempo_carrera_s a float (segundos); acepta número o hh:mm:ss.ccc. None si vacío o inválido."""
    if not s or not str(s).strip():
        return None
    t = str(s).strip().replace(",", ".")
    # Formato hh:mm:ss.ccc
    if ":" in t:
        parts = t.split(":")
        if len(parts) == 3:
            try:
                h, m, sec = int(parts[0]), int(parts[1]), float(parts[2])
                return h * 3600 + m * 60 + sec
            except ValueError:
                pass
    try:
        return float(t)
    except ValueError:
        return None


def _segundos_a_hhmmss(segundos) -> str:
    """Convierte segundos (float) a string hh:mm:ss.ccc."""
    if segundos is None:
        return ""
    try:
        s = float(segundos)
    except (TypeError, ValueError):
        return ""
    if s < 0 or s != s:
        return ""
    h = int(s // 3600)
    m = int((s % 3600) // 60)
    sec = s % 60
    return f"{h:02d}:{m:02d}:{sec:06.3f}"


def cargar_resultados(ruta: str) -> list:
    """Carga resultados_con_nadadores.csv y devuelve lista de dicts."""
    contenido = _leer_archivo_texto(ruta)
    r = csv.reader(io.StringIO(contenido))
    filas = []
    cabecera = None
    for row in r:
        if not row:
            continue
        if row[0] == "inicio_punto_cero":
            continue
        if row[0] == "posicion":
            cabecera = row
            continue
        if cabecera is None or len(row) < len(cabecera):
            continue
        filas.append(dict(zip(cabecera, row)))
    return filas


def _asignar_posiciones(registros: list, clave_tiempo: str = "tiempo_carrera_s") -> list:
    """Ordena por tiempo (asc) y asigna posicion 1, 2, 3... (sin tiempo al final)."""
    con_tiempo = [(r, _parse_tiempo(r.get(clave_tiempo, ""))) for r in registros]
    con_tiempo.sort(key=lambda x: (x[1] is None, x[1] if x[1] is not None else 0))
    for i, (r, _) in enumerate(con_tiempo, 1):
        r["_pos"] = i
    return [r for r, _ in con_tiempo]


def calcular_clasificaciones(registros: list) -> list:
    """Añade posicion_general y posicion_categoria_sexo (simultáneo categoría + sexo) a cada registro."""
    regs = [dict(r) for r in registros]
    for r in regs:
        r["_tiempo"] = _parse_tiempo(r.get("tiempo_carrera_s") or r.get("tiempo_carrera", ""))

    # Clasificación general
    ordenados = sorted(regs, key=lambda r: (r["_tiempo"] is None, r["_tiempo"] if r["_tiempo"] is not None else 0))
    for i, r in enumerate(ordenados, 1):
        r["posicion_general"] = i

    # Por categoría y sexo (simultáneo)
    grupos = {}
    for r in regs:
        cat = (r.get("categoria_nombre") or "").strip() or "Sin categoría"
        sexo = (r.get("genero") or "").strip() or "Sin dato"
        clave = (cat, sexo)
        if clave not in grupos:
            grupos[clave] = []
        grupos[clave].append(r)
    for clave, lista in grupos.items():
        lista_sort = sorted(lista, key=lambda r: (r["_tiempo"] is None, r["_tiempo"] if r["_tiempo"] is not None else 0))
        for i, r in enumerate(lista_sort, 1):
            r["posicion_categoria_sexo"] = i

    for r in regs:
        r.pop("_tiempo", None)
    return sorted(regs, key=lambda r: r["posicion_general"])


def guardar_csv(registros: list, ruta: str) -> None:
    """Guarda CSV con posición general, posición categoría+sexo y datos."""
    if not registros:
        return
    cabecera = [
        "posicion_general", "posicion_categoria_sexo",
        "nombre", "numero_corredor", "categoria_nombre", "genero", "distancia",
        "tiempo_carrera", "hora_llegada", "posicion", "epc", "antena", "rssi", "epc_en_planilla"
    ]
    with open(ruta, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cabecera, extrasaction="ignore")
        w.writeheader()
        for r in registros:
            row = {k: r.get(k, "") for k in cabecera}
            row["tiempo_carrera"] = _segundos_a_hhmmss(_parse_tiempo(r.get("tiempo_carrera_s", "")))
            w.writerow(row)
    print(f"✓ CSV guardado: {ruta}")


def guardar_excel(registros: list, ruta: str) -> None:
    """Guarda la clasificación en un libro Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        print("⚠ Para exportar Excel instala: pip install openpyxl")
        return
    if not registros:
        return
    cabecera = [
        "posicion_general", "posicion_categoria_sexo",
        "nombre", "numero_corredor", "categoria_nombre", "genero", "distancia",
        "tiempo_carrera", "hora_llegada", "posicion", "epc", "antena", "rssi", "epc_en_planilla"
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Clasificación"
    for c, col in enumerate(cabecera, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = Font(bold=True)
    for row_idx, r in enumerate(registros, 2):
        for c, col in enumerate(cabecera, 1):
            val = _segundos_a_hhmmss(_parse_tiempo(r.get("tiempo_carrera_s", ""))) if col == "tiempo_carrera" else r.get(col, "")
            ws.cell(row=row_idx, column=c, value=val)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14
    wb.save(ruta)
    print(f"✓ Excel guardado: {ruta}")


def guardar_pdf(registros: list, ruta: str) -> None:
    """Genera PDF con clasificación general y por categoría+sexo (simultáneo)."""
    try:
        from fpdf import FPDF
    except ImportError:
        print("⚠ Para generar el PDF instala: pip install fpdf2")
        return
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, _pdf_safe("Clasificacion por tiempo", 30), ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, _pdf_safe(f"Total participantes: {len(registros)}", 40), ln=True, align="C")
    pdf.ln(8)

    def tabla(pdf, titulo: str, filas: list, columnas: list, anchos: list):
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, _pdf_safe(titulo, 80), ln=True)
        pdf.set_font("Helvetica", "B", 8)
        for col, w in zip(columnas, anchos):
            pdf.cell(w, 6, _pdf_safe(str(col), 20), border=1)
        pdf.ln()
        pdf.set_font("Helvetica", "", 8)
        for row in filas:
            for k, w in zip(columnas, anchos):
                val = _pdf_safe(row.get(k, ""), 20)
                pdf.cell(w, 5, val, border=1)
            pdf.ln()
        pdf.ln(4)

    # General
    col_gen = ["posicion_general", "nombre", "categoria_nombre", "genero", "tiempo_carrera"]
    w_gen = [18, 55, 40, 28, 28]
    registros_pdf = []
    for r in registros:
        rp = dict(r)
        rp["tiempo_carrera"] = _segundos_a_hhmmss(_parse_tiempo(r.get("tiempo_carrera_s", "")))
        registros_pdf.append(rp)
    tabla(pdf, "1. Clasificacion general", registros_pdf, col_gen, w_gen)

    # Por categoría y sexo (simultáneo)
    grupos = {}
    for r in registros:
        cat = (r.get("categoria_nombre") or "").strip() or "Sin categoría"
        sexo = (r.get("genero") or "").strip() or "Sin dato"
        clave = (cat, sexo)
        if clave not in grupos:
            grupos[clave] = []
        grupos[clave].append(r)
    col_cs = ["posicion_categoria_sexo", "nombre", "tiempo_carrera"]
    w_cs = [28, 85, 40]
    for (cat, sexo) in sorted(grupos.keys()):
        lista = sorted(grupos[(cat, sexo)], key=lambda r: r["posicion_categoria_sexo"])
        lista_pdf = []
        for r in lista:
            rp = dict(r)
            rp["tiempo_carrera"] = _segundos_a_hhmmss(_parse_tiempo(r.get("tiempo_carrera_s", "")))
            lista_pdf.append(rp)
        titulo = f"2. Categoria y sexo: {cat} - {sexo}"
        tabla(pdf, titulo, lista_pdf, col_cs, w_cs)

    pdf.output(ruta)
    print(f"✓ PDF guardado: {ruta}")


def main(entrada: str = CSV_ENTRADA, salida_csv: str = CSV_SALIDA,
         salida_xlsx: str = XLSX_SALIDA, salida_pdf: str = PDF_SALIDA) -> bool:
    """Carga resultados, calcula clasificaciones y genera CSV, Excel y PDF."""
    try:
        registros = cargar_resultados(entrada)
    except FileNotFoundError:
        print(f"⚠ No se encontró {entrada}. Ejecuta antes: python cruzar_resultados.py")
        return False
    if not registros:
        print("⚠ No hay filas de resultados en el CSV.")
        return False
    con_pos = calcular_clasificaciones(registros)
    guardar_csv(con_pos, salida_csv)
    guardar_excel(con_pos, salida_xlsx)
    guardar_pdf(con_pos, salida_pdf)
    return True


if __name__ == "__main__":
    entrada = sys.argv[1] if len(sys.argv) > 1 else CSV_ENTRADA
    if not main(entrada=entrada):
        sys.exit(1)
