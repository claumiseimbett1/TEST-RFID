#!/usr/bin/env python3
"""
Genera una planilla de inscripción en Excel (nombre, categoria, genero, club, tiempo_inscripcion).

Uso:
  python generar_planilla_inscripcion.py [salida.xlsx] [-n filas]
  Ejemplo: python generar_planilla_inscripcion.py planilla_inscripcion.xlsx -n 150

Exportar a CSV: python generar_planilla_inscripcion.py --export planilla_inscripcion.xlsx inscripcion.csv

Generar nombres_nadadores.csv desde la planilla:
  python generar_planilla_inscripcion.py --nombres planilla_inscripcion.xlsx nombres_nadadores.csv
  Opcional: --tags tags_para_registro.csv para asignar EPC por coincidencia nombre/categoria/genero.
"""
import argparse
import csv
import sys
from typing import List, Optional

COLUMNAS = ["nombre", "categoria", "genero", "club", "tiempo_inscripcion"]
COLUMNAS_NOMBRES_NADADORES = ["epc", "nombre", "categoria", "sexo"]
SALIDA_DEFECTO = "planilla_inscripcion.xlsx"
FILAS_DEFECTO = 100
NOMBRES_NADADORES_CSV = "nombres_nadadores.csv"


def _normalizar_para_clave(s: str) -> str:
    """Normaliza string para emparejar (quita acentos, minúsculas, espacios)."""
    if not s:
        return ""
    t = str(s).strip().lower()
    for a, b in [
        ("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n"), ("ü", "u")
    ]:
        t = t.replace(a, b)
    return " ".join(t.split())


def _cargar_inscripcion(ruta: str) -> List[dict]:
    """Carga planilla de inscripción desde Excel o CSV. Devuelve lista de dicts con nombre, categoria, genero."""
    ruta_lower = ruta.lower()
    if ruta_lower.endswith(".xlsx") or ruta_lower.endswith(".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("Para leer Excel instala: pip install openpyxl")
        wb = load_workbook(ruta, read_only=True, data_only=True)
        ws = wb.active
        filas = list(ws.iter_rows(values_only=True))
        wb.close()
        if not filas:
            return []
        cabecera = [str(c).strip().lower() if c is not None else "" for c in filas[0]]
        # Mapear a nombre, categoria, genero (aceptar variantes)
        key_n = "nombre" if "nombre" in cabecera else (cabecera[0] if cabecera else "")
        key_c = "categoria" if "categoria" in cabecera else "categoria_nombre" if "categoria_nombre" in cabecera else (cabecera[1] if len(cabecera) > 1 else "")
        key_g = "genero" if "genero" in cabecera else "sexo" if "sexo" in cabecera else (cabecera[2] if len(cabecera) > 2 else "")
        idx_n = cabecera.index(key_n) if key_n in cabecera else 0
        idx_c = cabecera.index(key_c) if key_c in cabecera else 1
        idx_g = cabecera.index(key_g) if key_g in cabecera else 2
        out = []
        for row in filas[1:]:
            if not row:
                continue
            vals = list(row) if row else []
            nombre = str(vals[idx_n]).strip() if idx_n < len(vals) and vals[idx_n] is not None else ""
            categoria = str(vals[idx_c]).strip() if idx_c < len(vals) and vals[idx_c] is not None else ""
            genero = str(vals[idx_g]).strip() if idx_g < len(vals) and vals[idx_g] is not None else ""
            if not nombre and not categoria and not genero:
                continue
            out.append({"nombre": nombre, "categoria": categoria, "genero": genero})
        return out
    # CSV
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            with open(ruta, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return []
                keys = {k.strip().lstrip("\ufeff").lower(): k for k in reader.fieldnames}
                col_n = keys.get("nombre") or (list(keys.values())[0] if keys else "")
                col_c = keys.get("categoria") or keys.get("categoria_nombre") or (list(keys.values())[1] if len(keys) > 1 else "")
                col_g = keys.get("genero") or keys.get("sexo") or (list(keys.values())[2] if len(keys) > 2 else "")
                out = []
                for row in reader:
                    nombre = (row.get(col_n) or "").strip()
                    categoria = (row.get(col_c) or "").strip()
                    genero = (row.get(col_g) or "").strip()
                    if not nombre and not categoria and not genero:
                        continue
                    out.append({"nombre": nombre, "categoria": categoria, "genero": genero})
                return out
        except (UnicodeDecodeError, KeyError, FileNotFoundError):
            continue
    return []


def _cargar_lista_epcs_desde_tags(tags_path: str) -> List[str]:
    """
    Lee tags_para_registro.csv y devuelve la lista de EPCs en orden de fila
    (columna epc_formateado o primera columna).
    """
    epcs: List[str] = []
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            with open(tags_path, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                fn = reader.fieldnames or []
                keys = {k.strip().lstrip("\ufeff").lower(): k for k in fn}
                col_epc = keys.get("epc") or keys.get("epc_formateado") or (fn[0] if fn else "")
                if not col_epc:
                    break
                for row in reader:
                    epc = (row.get(col_epc) or "").strip()
                    epcs.append(epc)
                return epcs
        except (UnicodeDecodeError, FileNotFoundError):
            continue
    return epcs


def _cargar_epc_por_nombre_categoria_genero(tags_path: str) -> dict:
    """Construye clave (nombre_norm, categoria_norm, genero_norm) -> epc desde tags (solo si tiene columna nombre)."""
    lookup = {}
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            with open(tags_path, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                fn = reader.fieldnames or []
                keys = {k.strip().lstrip("\ufeff").lower(): k for k in fn}
                col_epc = keys.get("epc") or keys.get("epc_formateado") or (fn[0] if fn else "")
                col_n = keys.get("nombre") or keys.get("nombre_nadador")
                if not col_epc or not col_n:
                    break
                col_c = keys.get("categoria") or keys.get("categoria_nombre") or (fn[2] if len(fn) > 2 else "")
                col_g = keys.get("genero") or keys.get("sexo") or (fn[3] if len(fn) > 3 else "")
                for row in reader:
                    epc = (row.get(col_epc) or "").strip()
                    if not epc:
                        continue
                    nombre = (row.get(col_n) or "").strip()
                    categoria = (row.get(col_c) or "").strip() if col_c else ""
                    genero = (row.get(col_g) or "").strip() if col_g else ""
                    clave = (_normalizar_para_clave(nombre), _normalizar_para_clave(categoria), _normalizar_para_clave(genero))
                    if clave not in lookup:
                        lookup[clave] = epc
                return lookup
        except (UnicodeDecodeError, FileNotFoundError):
            continue
    return lookup


def generar_nombres_nadadores_desde_inscripcion(
    inscripcion_path: str,
    salida_path: str = NOMBRES_NADADORES_CSV,
    planilla_tags_path: Optional[str] = None,
) -> bool:
    """
    Lee la planilla de inscripción (Excel o CSV con nombre, categoria, genero, ...)
    y genera nombres_nadadores.csv con columnas epc, nombre, categoria, sexo.

    Los EPC se leen siempre desde tags_para_registro (planilla_tags_path) por orden de fila:
    fila 1 de inscripción = fila 1 de tags, etc. Si no se pasa planilla_tags_path, EPC queda vacío.
    """
    try:
        inscritos = _cargar_inscripcion(inscripcion_path)
    except FileNotFoundError:
        print(f"No se encontró: {inscripcion_path}")
        return False
    except RuntimeError as e:
        print(e)
        return False

    if not inscritos:
        print("La planilla de inscripción no tiene filas con datos.")
        return False

    # EPCs desde tags_para_registro en orden de fila (columna epc_formateado)
    lista_epcs: List[str] = []
    if planilla_tags_path:
        try:
            lista_epcs = _cargar_lista_epcs_desde_tags(planilla_tags_path)
            if lista_epcs:
                print(f"  EPCs leídos desde {planilla_tags_path} ({len(lista_epcs)} tags, asignación por orden de fila)")
            else:
                print(f"  No se encontraron EPCs en {planilla_tags_path}")
        except FileNotFoundError:
            print(f"  No se encontró: {planilla_tags_path}. EPC quedará vacío.")

    with open(salida_path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNAS_NOMBRES_NADADORES)
        w.writeheader()
        for i, r in enumerate(inscritos):
            epc = lista_epcs[i] if i < len(lista_epcs) else ""
            w.writerow({
                "epc": epc,
                "nombre": r["nombre"],
                "categoria": r["categoria"],
                "sexo": r["genero"],
            })

    print(f"Generado: {salida_path} ({len(inscritos)} filas)")
    return True


def generar_excel(ruta: str, num_filas: int = FILAS_DEFECTO) -> bool:
    """Crea un Excel con cabecera nombre, categoria, genero, club, tiempo_inscripcion y filas vacías."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("Para generar Excel instala: pip install openpyxl")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Inscripcion"

    # Cabecera
    for c, col in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="medium"),
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
        )

    # Filas vacías para completar
    for r in range(2, 2 + num_filas):
        for c in range(1, len(COLUMNAS) + 1):
            ws.cell(row=r, column=c, value="")

    # Validación de lista para genero
    dv_genero = DataValidation(
        type="list",
        formula1='"Masculino,Femenino"',
        allow_blank=True,
    )
    dv_genero.error = "Elija Masculino o Femenino"
    dv_genero.errorTitle = "Género"
    ws.add_data_validation(dv_genero)
    dv_genero.add(f"C2:C{1 + num_filas}")

    # Anchos de columna
    anchos = [45, 25, 14, 30, 20]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Congelar primera fila
    ws.freeze_panes = "A2"

    wb.save(ruta)
    print(f"Planilla creada: {ruta}")
    print(f"  Columnas: {', '.join(COLUMNAS)} ({num_filas} filas para completar)")
    return True


def exportar_a_csv(excel_ruta: str, csv_ruta: str) -> bool:
    """Exporta la planilla Excel a CSV (UTF-8)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Para exportar instala: pip install openpyxl")
        return False
    try:
        wb = load_workbook(excel_ruta, read_only=True, data_only=True)
        ws = wb.active
        filas = list(ws.iter_rows(values_only=True))
        wb.close()
    except FileNotFoundError:
        print(f"No se encontró: {excel_ruta}")
        return False
    if not filas:
        print("El archivo Excel está vacío.")
        return False
    cabecera = [str(c).strip() if c is not None else "" for c in filas[0]]
    with open(csv_ruta, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(cabecera)
        for row in filas[1:]:
            w.writerow([str(v).strip() if v is not None else "" for v in row])
    print(f"Exportado: {csv_ruta} (desde {excel_ruta})")
    return True


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Genera planilla de inscripción Excel (nombre, categoria, genero, club, tiempo_inscripcion)"
    )
    parser.add_argument(
        "salida",
        nargs="?",
        default=SALIDA_DEFECTO,
        help=f"Archivo Excel de salida (default: {SALIDA_DEFECTO})",
    )
    parser.add_argument(
        "-n",
        "--filas",
        type=int,
        default=FILAS_DEFECTO,
        help=f"Filas vacías a crear (default: {FILAS_DEFECTO})",
    )
    parser.add_argument(
        "--export",
        nargs=2,
        metavar=("EXCEL", "CSV"),
        help="Exportar Excel a CSV: --export planilla_inscripcion.xlsx inscripcion.csv",
    )
    parser.add_argument(
        "--nombres",
        nargs="+",
        metavar=("INSCRIPCION", "SALIDA"),
        help="Generar nombres_nadadores.csv desde planilla: --nombres planilla.xlsx [nombres_nadadores.csv]",
    )
    parser.add_argument(
        "--tags",
        default=None,
        metavar="TAGS_CSV",
        help="Planilla de tags (p. ej. tags_para_registro.csv) para asignar EPC por nombre/categoria/genero",
    )
    args = parser.parse_args()

    if args.export:
        excel_path, csv_path = args.export
        return 0 if exportar_a_csv(excel_path, csv_path) else 1

    if args.nombres:
        inscripcion_path = args.nombres[0]
        salida_path = args.nombres[1] if len(args.nombres) > 1 else NOMBRES_NADADORES_CSV
        return 0 if generar_nombres_nadadores_desde_inscripcion(
            inscripcion_path, salida_path=salida_path, planilla_tags_path=args.tags
        ) else 1

    num_filas = args.filas if args.filas > 0 else FILAS_DEFECTO
    return 0 if generar_excel(args.salida, num_filas) else 1


if __name__ == "__main__":
    sys.exit(main())
